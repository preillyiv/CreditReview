"""
Excel export with formulas.

Generates an Excel workbook with:
- Sheet 1: Raw Values with source citations
- Sheet 2: Calculated Metrics with actual Excel formulas
- Sheet 3: Ratios with actual Excel formulas
- Sheet 4: Audit Log with calculation steps

All calculated cells use real Excel formulas, so users can:
- See exactly how values are derived
- Modify inputs and see calculations update
- Audit the entire calculation chain
"""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.models.extraction import (
    ExtractionSession,
    CalculationStep,
    METRIC_DISPLAY_NAMES,
    INCOME_STATEMENT_ITEMS,
    BALANCE_SHEET_ITEMS,
    CASH_FLOW_ITEMS,
)
from src.calculators.metrics import FinancialMetrics
from src.calculators.ratios import FinancialRatios
from src.calculators.verification import VerificationResult


# Style constants
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
POSITIVE_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
NEGATIVE_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FORMULA_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def _apply_header_style(cell):
    """Apply header styling to a cell."""
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = THIN_BORDER


def _apply_data_style(cell, is_formula=False):
    """Apply data cell styling."""
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal="right")
    if is_formula:
        cell.fill = FORMULA_FILL


def _format_number(value: float, is_currency: bool = True, is_percentage: bool = False) -> str:
    """Format a number for display."""
    if is_percentage:
        return f"{value:.2%}"
    if is_currency:
        if abs(value) >= 1e9:
            return f"${value/1e9:,.2f}B"
        elif abs(value) >= 1e6:
            return f"${value/1e6:,.2f}M"
        else:
            return f"${value:,.0f}"
    return f"{value:,.2f}"


def generate_excel_report(
    session: ExtractionSession,
    metrics: FinancialMetrics,
    ratios: FinancialRatios,
    calculation_steps: list[CalculationStep],
    verification: VerificationResult = None,
) -> io.BytesIO:
    """
    Generate an Excel workbook with financial data and formulas.

    Args:
        session: The extraction session with raw values
        metrics: Calculated financial metrics
        ratios: Calculated financial ratios
        calculation_steps: Audit trail of calculations
        verification: Optional verification results

    Returns:
        BytesIO buffer containing the Excel file
    """
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Create sheets
    _create_raw_values_sheet(wb, session)
    _create_statement_sheet(wb, session, "Income Statement", INCOME_STATEMENT_ITEMS)
    _create_statement_sheet(wb, session, "Balance Sheet", BALANCE_SHEET_ITEMS)
    _create_statement_sheet(wb, session, "Cash Flow", CASH_FLOW_ITEMS)
    _create_metrics_sheet(wb, session, metrics)
    _create_ratios_sheet(wb, session, ratios, metrics)
    if verification:
        _create_verification_sheet(wb, verification)
    _create_audit_sheet(wb, calculation_steps)

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer


def _create_raw_values_sheet(wb: Workbook, session: ExtractionSession):
    """Create the Raw Values sheet with source citations."""
    ws = wb.create_sheet("Raw Values")

    # Add unit note if not in dollars
    if session.unit and session.unit.lower() != "dollars":
        note_cell = ws.cell(row=1, column=1, value=f"All figures are in {session.unit}")
        note_cell.font = Font(italic=True, color="808080", size=10)
        ws.row_dimensions[1].height = 20

    # Headers (starting at row 2 or 1 depending on unit note)
    header_row = 2 if (session.unit and session.unit.lower() != "dollars") else 1
    headers = ["Metric", "Current Value", "Prior Value", "XBRL Concept", "Filing Date", "SEC Link"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        _apply_header_style(cell)

    # Data rows
    row = header_row + 1
    for metric_key, ev in session.raw_values.items():
        ws.cell(row=row, column=1, value=ev.display_name)
        ws.cell(row=row, column=2, value=ev.value)
        ws.cell(row=row, column=3, value=ev.value_prior)

        if ev.citation:
            ws.cell(row=row, column=4, value=ev.citation.xbrl_concept)
            ws.cell(row=row, column=5, value=ev.citation.filing_date)
            ws.cell(row=row, column=6, value=ev.citation.filing_url)

        # Apply number format
        ws.cell(row=row, column=2).number_format = '#,##0'
        ws.cell(row=row, column=3).number_format = '#,##0'

        row += 1

    # Add company info header
    row += 2
    ws.cell(row=row, column=1, value="Company Information").font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=1, value="Ticker")
    ws.cell(row=row, column=2, value=session.ticker)
    row += 1
    ws.cell(row=row, column=1, value="Company Name")
    ws.cell(row=row, column=2, value=session.company_name)
    row += 1
    ws.cell(row=row, column=1, value="CIK")
    ws.cell(row=row, column=2, value=session.cik)
    row += 1
    ws.cell(row=row, column=1, value="Fiscal Year End")
    ws.cell(row=row, column=2, value=session.fiscal_year_end)
    row += 1
    ws.cell(row=row, column=1, value="Prior Fiscal Year End")
    ws.cell(row=row, column=2, value=session.fiscal_year_end_prior)

    # Auto-fit columns
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 20


def _create_metrics_sheet(wb: Workbook, session: ExtractionSession, metrics: FinancialMetrics):
    """Create the Calculated Metrics sheet with Excel formulas."""
    ws = wb.create_sheet("Calculated Metrics")

    # Create a mapping of raw values to row numbers for formulas
    # First, list the raw value rows
    raw_row_map = {}

    # Headers
    headers = ["Metric", "Current", "Prior", "YoY Delta", "Formula"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        _apply_header_style(cell)

    # Raw values section
    ws.cell(row=3, column=1, value="RAW VALUES (from XBRL)").font = Font(bold=True, color="1F4E79")

    row = 4
    raw_metrics = ["revenue", "cost_of_revenue", "gross_profit", "operating_income",
                   "depreciation_amortization", "net_income", "stockholders_equity",
                   "intangible_assets", "goodwill", "cash", "stock_compensation"]

    for metric_key in raw_metrics:
        if metric_key in session.raw_values:
            ev = session.raw_values[metric_key]
            raw_row_map[metric_key] = row
            ws.cell(row=row, column=1, value=ev.display_name)
            ws.cell(row=row, column=2, value=ev.value)
            ws.cell(row=row, column=3, value=ev.value_prior)
            # Delta formula
            ws.cell(row=row, column=4, value=f"=B{row}-C{row}")
            _apply_data_style(ws.cell(row=row, column=4), is_formula=True)
            ws.cell(row=row, column=5, value="Raw value from SEC filing")

            # Number formats
            ws.cell(row=row, column=2).number_format = '#,##0'
            ws.cell(row=row, column=3).number_format = '#,##0'
            ws.cell(row=row, column=4).number_format = '#,##0'

            row += 1

    # Calculated metrics section
    row += 1
    ws.cell(row=row, column=1, value="CALCULATED METRICS").font = Font(bold=True, color="1F4E79")
    row += 1

    # Gross Margin
    ws.cell(row=row, column=1, value="Gross Profit Margin")
    if "gross_profit" in raw_row_map and "revenue" in raw_row_map:
        gp_row = raw_row_map["gross_profit"]
        rev_row = raw_row_map["revenue"]
        ws.cell(row=row, column=2, value=f"=B{gp_row}/B{rev_row}")
        ws.cell(row=row, column=3, value=f"=C{gp_row}/C{rev_row}")
        _apply_data_style(ws.cell(row=row, column=2), is_formula=True)
        _apply_data_style(ws.cell(row=row, column=3), is_formula=True)
    else:
        ws.cell(row=row, column=2, value=metrics.gross_profit_margin)
        ws.cell(row=row, column=3, value=metrics.gross_profit_margin_prior)
    ws.cell(row=row, column=4, value=f"=B{row}-C{row}")
    _apply_data_style(ws.cell(row=row, column=4), is_formula=True)
    ws.cell(row=row, column=5, value="Gross Profit / Revenue")
    ws.cell(row=row, column=2).number_format = '0.00%'
    ws.cell(row=row, column=3).number_format = '0.00%'
    ws.cell(row=row, column=4).number_format = '0.00%'
    gross_margin_row = row
    row += 1

    # Operating Margin
    ws.cell(row=row, column=1, value="Operating Income Margin")
    if "operating_income" in raw_row_map and "revenue" in raw_row_map:
        oi_row = raw_row_map["operating_income"]
        rev_row = raw_row_map["revenue"]
        ws.cell(row=row, column=2, value=f"=B{oi_row}/B{rev_row}")
        ws.cell(row=row, column=3, value=f"=C{oi_row}/C{rev_row}")
        _apply_data_style(ws.cell(row=row, column=2), is_formula=True)
        _apply_data_style(ws.cell(row=row, column=3), is_formula=True)
    else:
        ws.cell(row=row, column=2, value=metrics.operating_income_margin)
        ws.cell(row=row, column=3, value=metrics.operating_income_margin_prior)
    ws.cell(row=row, column=4, value=f"=B{row}-C{row}")
    _apply_data_style(ws.cell(row=row, column=4), is_formula=True)
    ws.cell(row=row, column=5, value="Operating Income / Revenue")
    ws.cell(row=row, column=2).number_format = '0.00%'
    ws.cell(row=row, column=3).number_format = '0.00%'
    ws.cell(row=row, column=4).number_format = '0.00%'
    row += 1

    # EBITDA
    ws.cell(row=row, column=1, value="EBITDA")
    if "operating_income" in raw_row_map and "depreciation_amortization" in raw_row_map:
        oi_row = raw_row_map["operating_income"]
        da_row = raw_row_map["depreciation_amortization"]
        ws.cell(row=row, column=2, value=f"=B{oi_row}+B{da_row}")
        ws.cell(row=row, column=3, value=f"=C{oi_row}+C{da_row}")
        _apply_data_style(ws.cell(row=row, column=2), is_formula=True)
        _apply_data_style(ws.cell(row=row, column=3), is_formula=True)
    else:
        ws.cell(row=row, column=2, value=metrics.ebitda)
        ws.cell(row=row, column=3, value=metrics.ebitda_prior)
    ws.cell(row=row, column=4, value=f"=B{row}-C{row}")
    _apply_data_style(ws.cell(row=row, column=4), is_formula=True)
    ws.cell(row=row, column=5, value="Operating Income + D&A")
    ws.cell(row=row, column=2).number_format = '#,##0'
    ws.cell(row=row, column=3).number_format = '#,##0'
    ws.cell(row=row, column=4).number_format = '#,##0'
    ebitda_row = row
    row += 1

    # EBITDA Margin
    ws.cell(row=row, column=1, value="EBITDA Margin")
    if "revenue" in raw_row_map:
        rev_row = raw_row_map["revenue"]
        ws.cell(row=row, column=2, value=f"=B{ebitda_row}/B{rev_row}")
        ws.cell(row=row, column=3, value=f"=C{ebitda_row}/C{rev_row}")
        _apply_data_style(ws.cell(row=row, column=2), is_formula=True)
        _apply_data_style(ws.cell(row=row, column=3), is_formula=True)
    else:
        ws.cell(row=row, column=2, value=metrics.ebitda_margin)
        ws.cell(row=row, column=3, value=metrics.ebitda_margin_prior)
    ws.cell(row=row, column=4, value=f"=B{row}-C{row}")
    _apply_data_style(ws.cell(row=row, column=4), is_formula=True)
    ws.cell(row=row, column=5, value="EBITDA / Revenue")
    ws.cell(row=row, column=2).number_format = '0.00%'
    ws.cell(row=row, column=3).number_format = '0.00%'
    ws.cell(row=row, column=4).number_format = '0.00%'
    row += 1

    # Adjusted EBITDA
    ws.cell(row=row, column=1, value="Adjusted EBITDA")
    if "stock_compensation" in raw_row_map:
        sbc_row = raw_row_map["stock_compensation"]
        ws.cell(row=row, column=2, value=f"=B{ebitda_row}+B{sbc_row}")
        ws.cell(row=row, column=3, value=f"=C{ebitda_row}+C{sbc_row}")
        _apply_data_style(ws.cell(row=row, column=2), is_formula=True)
        _apply_data_style(ws.cell(row=row, column=3), is_formula=True)
    else:
        ws.cell(row=row, column=2, value=metrics.adjusted_ebitda)
        ws.cell(row=row, column=3, value=metrics.adjusted_ebitda_prior)
    ws.cell(row=row, column=4, value=f"=B{row}-C{row}")
    _apply_data_style(ws.cell(row=row, column=4), is_formula=True)
    ws.cell(row=row, column=5, value="EBITDA + Stock Compensation")
    ws.cell(row=row, column=2).number_format = '#,##0'
    ws.cell(row=row, column=3).number_format = '#,##0'
    ws.cell(row=row, column=4).number_format = '#,##0'
    adj_ebitda_row = row
    row += 1

    # Net Margin
    ws.cell(row=row, column=1, value="Net Income Margin")
    if "net_income" in raw_row_map and "revenue" in raw_row_map:
        ni_row = raw_row_map["net_income"]
        rev_row = raw_row_map["revenue"]
        ws.cell(row=row, column=2, value=f"=B{ni_row}/B{rev_row}")
        ws.cell(row=row, column=3, value=f"=C{ni_row}/C{rev_row}")
        _apply_data_style(ws.cell(row=row, column=2), is_formula=True)
        _apply_data_style(ws.cell(row=row, column=3), is_formula=True)
    else:
        ws.cell(row=row, column=2, value=metrics.net_income_margin)
        ws.cell(row=row, column=3, value=metrics.net_income_margin_prior)
    ws.cell(row=row, column=4, value=f"=B{row}-C{row}")
    _apply_data_style(ws.cell(row=row, column=4), is_formula=True)
    ws.cell(row=row, column=5, value="Net Income / Revenue")
    ws.cell(row=row, column=2).number_format = '0.00%'
    ws.cell(row=row, column=3).number_format = '0.00%'
    ws.cell(row=row, column=4).number_format = '0.00%'
    row += 1

    # Tangible Net Worth
    ws.cell(row=row, column=1, value="Tangible Net Worth")
    if all(k in raw_row_map for k in ["stockholders_equity", "intangible_assets", "goodwill"]):
        se_row = raw_row_map["stockholders_equity"]
        ia_row = raw_row_map["intangible_assets"]
        gw_row = raw_row_map["goodwill"]
        ws.cell(row=row, column=2, value=f"=B{se_row}-B{ia_row}-B{gw_row}")
        ws.cell(row=row, column=3, value=f"=C{se_row}-C{ia_row}-C{gw_row}")
        _apply_data_style(ws.cell(row=row, column=2), is_formula=True)
        _apply_data_style(ws.cell(row=row, column=3), is_formula=True)
    else:
        ws.cell(row=row, column=2, value=metrics.tangible_net_worth)
        ws.cell(row=row, column=3, value=metrics.tangible_net_worth_prior)
    ws.cell(row=row, column=4, value=f"=B{row}-C{row}")
    _apply_data_style(ws.cell(row=row, column=4), is_formula=True)
    ws.cell(row=row, column=5, value="Equity - Intangibles - Goodwill")
    ws.cell(row=row, column=2).number_format = '#,##0'
    ws.cell(row=row, column=3).number_format = '#,##0'
    ws.cell(row=row, column=4).number_format = '#,##0'

    # Auto-fit columns
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 35


def _create_ratios_sheet(
    wb: Workbook,
    session: ExtractionSession,
    ratios: FinancialRatios,
    metrics: FinancialMetrics
):
    """Create the Ratios sheet with Excel formulas."""
    ws = wb.create_sheet("Ratios")

    # Create a mapping of raw values to row numbers
    raw_row_map = {}

    # Headers
    headers = ["Ratio", "Current", "Prior", "YoY Delta", "Formula"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        _apply_header_style(cell)

    # Raw values needed for ratios
    ws.cell(row=3, column=1, value="RAW VALUES (from XBRL)").font = Font(bold=True, color="1F4E79")

    row = 4
    raw_metrics = ["current_assets", "current_liabilities", "cash", "total_debt",
                   "stockholders_equity", "interest_expense", "accounts_receivable",
                   "revenue", "net_income", "total_assets"]

    for metric_key in raw_metrics:
        if metric_key in session.raw_values:
            ev = session.raw_values[metric_key]
            raw_row_map[metric_key] = row
            ws.cell(row=row, column=1, value=ev.display_name)
            ws.cell(row=row, column=2, value=ev.value)
            ws.cell(row=row, column=3, value=ev.value_prior)
            ws.cell(row=row, column=4, value=f"=B{row}-C{row}")
            _apply_data_style(ws.cell(row=row, column=4), is_formula=True)

            ws.cell(row=row, column=2).number_format = '#,##0'
            ws.cell(row=row, column=3).number_format = '#,##0'
            ws.cell(row=row, column=4).number_format = '#,##0'

            row += 1

    # Add EBITDA from metrics calculation
    ws.cell(row=row, column=1, value="EBITDA (calculated)")
    ws.cell(row=row, column=2, value=metrics.ebitda)
    ws.cell(row=row, column=3, value=metrics.ebitda_prior)
    ws.cell(row=row, column=4, value=f"=B{row}-C{row}")
    ws.cell(row=row, column=2).number_format = '#,##0'
    ws.cell(row=row, column=3).number_format = '#,##0'
    ws.cell(row=row, column=4).number_format = '#,##0'
    ebitda_row = row
    row += 1

    # Calculated ratios section
    row += 1
    ws.cell(row=row, column=1, value="CALCULATED RATIOS").font = Font(bold=True, color="1F4E79")
    row += 1

    # Current Ratio
    ws.cell(row=row, column=1, value="Current Ratio")
    if "current_assets" in raw_row_map and "current_liabilities" in raw_row_map:
        ca_row = raw_row_map["current_assets"]
        cl_row = raw_row_map["current_liabilities"]
        ws.cell(row=row, column=2, value=f"=B{ca_row}/B{cl_row}")
        ws.cell(row=row, column=3, value=f"=C{ca_row}/C{cl_row}")
        _apply_data_style(ws.cell(row=row, column=2), is_formula=True)
        _apply_data_style(ws.cell(row=row, column=3), is_formula=True)
    else:
        ws.cell(row=row, column=2, value=ratios.current_ratio)
        ws.cell(row=row, column=3, value=ratios.current_ratio_prior)
    ws.cell(row=row, column=4, value=f"=B{row}-C{row}")
    _apply_data_style(ws.cell(row=row, column=4), is_formula=True)
    ws.cell(row=row, column=5, value="Current Assets / Current Liabilities")
    ws.cell(row=row, column=2).number_format = '0.00'
    ws.cell(row=row, column=3).number_format = '0.00'
    ws.cell(row=row, column=4).number_format = '+0.00;-0.00'
    row += 1

    # Cash Ratio
    ws.cell(row=row, column=1, value="Cash Ratio")
    if "cash" in raw_row_map and "current_liabilities" in raw_row_map:
        cash_row = raw_row_map["cash"]
        cl_row = raw_row_map["current_liabilities"]
        ws.cell(row=row, column=2, value=f"=B{cash_row}/B{cl_row}")
        ws.cell(row=row, column=3, value=f"=C{cash_row}/C{cl_row}")
        _apply_data_style(ws.cell(row=row, column=2), is_formula=True)
        _apply_data_style(ws.cell(row=row, column=3), is_formula=True)
    else:
        ws.cell(row=row, column=2, value=ratios.cash_ratio)
        ws.cell(row=row, column=3, value=ratios.cash_ratio_prior)
    ws.cell(row=row, column=4, value=f"=B{row}-C{row}")
    _apply_data_style(ws.cell(row=row, column=4), is_formula=True)
    ws.cell(row=row, column=5, value="Cash / Current Liabilities")
    ws.cell(row=row, column=2).number_format = '0.00'
    ws.cell(row=row, column=3).number_format = '0.00'
    ws.cell(row=row, column=4).number_format = '+0.00;-0.00'
    row += 1

    # Debt-to-Equity
    ws.cell(row=row, column=1, value="Debt-to-Equity")
    if "total_debt" in raw_row_map and "stockholders_equity" in raw_row_map:
        td_row = raw_row_map["total_debt"]
        se_row = raw_row_map["stockholders_equity"]
        ws.cell(row=row, column=2, value=f"=B{td_row}/B{se_row}")
        ws.cell(row=row, column=3, value=f"=C{td_row}/C{se_row}")
        _apply_data_style(ws.cell(row=row, column=2), is_formula=True)
        _apply_data_style(ws.cell(row=row, column=3), is_formula=True)
    else:
        ws.cell(row=row, column=2, value=ratios.debt_to_equity)
        ws.cell(row=row, column=3, value=ratios.debt_to_equity_prior)
    ws.cell(row=row, column=4, value=f"=B{row}-C{row}")
    _apply_data_style(ws.cell(row=row, column=4), is_formula=True)
    ws.cell(row=row, column=5, value="Total Debt / Stockholders' Equity")
    ws.cell(row=row, column=2).number_format = '0.00'
    ws.cell(row=row, column=3).number_format = '0.00'
    ws.cell(row=row, column=4).number_format = '+0.00;-0.00'
    row += 1

    # EBITDA Interest Coverage
    ws.cell(row=row, column=1, value="EBITDA Interest Coverage")
    if "interest_expense" in raw_row_map:
        ie_row = raw_row_map["interest_expense"]
        ws.cell(row=row, column=2, value=f"=B{ebitda_row}/B{ie_row}")
        ws.cell(row=row, column=3, value=f"=C{ebitda_row}/C{ie_row}")
        _apply_data_style(ws.cell(row=row, column=2), is_formula=True)
        _apply_data_style(ws.cell(row=row, column=3), is_formula=True)
    else:
        ws.cell(row=row, column=2, value=ratios.ebitda_interest_coverage)
        ws.cell(row=row, column=3, value=ratios.ebitda_interest_coverage_prior)
    ws.cell(row=row, column=4, value=f"=B{row}-C{row}")
    _apply_data_style(ws.cell(row=row, column=4), is_formula=True)
    ws.cell(row=row, column=5, value="EBITDA / Interest Expense")
    ws.cell(row=row, column=2).number_format = '0.00'
    ws.cell(row=row, column=3).number_format = '0.00'
    ws.cell(row=row, column=4).number_format = '+0.00;-0.00'
    row += 1

    # Net Debt / EBITDA
    ws.cell(row=row, column=1, value="Net Debt / EBITDA")
    if "total_debt" in raw_row_map and "cash" in raw_row_map:
        td_row = raw_row_map["total_debt"]
        cash_row = raw_row_map["cash"]
        ws.cell(row=row, column=2, value=f"=(B{td_row}-B{cash_row})/B{ebitda_row}")
        ws.cell(row=row, column=3, value=f"=(C{td_row}-C{cash_row})/C{ebitda_row}")
        _apply_data_style(ws.cell(row=row, column=2), is_formula=True)
        _apply_data_style(ws.cell(row=row, column=3), is_formula=True)
    else:
        ws.cell(row=row, column=2, value=ratios.net_debt_to_ebitda)
        ws.cell(row=row, column=3, value=ratios.net_debt_to_ebitda_prior)
    ws.cell(row=row, column=4, value=f"=B{row}-C{row}")
    _apply_data_style(ws.cell(row=row, column=4), is_formula=True)
    ws.cell(row=row, column=5, value="(Total Debt - Cash) / EBITDA")
    ws.cell(row=row, column=2).number_format = '0.00'
    ws.cell(row=row, column=3).number_format = '0.00'
    ws.cell(row=row, column=4).number_format = '+0.00;-0.00'
    row += 1

    # Days Sales Outstanding
    ws.cell(row=row, column=1, value="Days Sales Outstanding")
    if "accounts_receivable" in raw_row_map and "revenue" in raw_row_map:
        ar_row = raw_row_map["accounts_receivable"]
        rev_row = raw_row_map["revenue"]
        ws.cell(row=row, column=2, value=f"=(B{ar_row}/B{rev_row})*365")
        ws.cell(row=row, column=3, value=f"=(C{ar_row}/C{rev_row})*365")
        _apply_data_style(ws.cell(row=row, column=2), is_formula=True)
        _apply_data_style(ws.cell(row=row, column=3), is_formula=True)
    else:
        ws.cell(row=row, column=2, value=ratios.days_sales_outstanding)
        ws.cell(row=row, column=3, value=ratios.days_sales_outstanding_prior)
    ws.cell(row=row, column=4, value=f"=B{row}-C{row}")
    _apply_data_style(ws.cell(row=row, column=4), is_formula=True)
    ws.cell(row=row, column=5, value="(A/R / Revenue) × 365")
    ws.cell(row=row, column=2).number_format = '0.0'
    ws.cell(row=row, column=3).number_format = '0.0'
    ws.cell(row=row, column=4).number_format = '+0.0;-0.0'
    row += 1

    # Working Capital
    ws.cell(row=row, column=1, value="Working Capital")
    if "current_assets" in raw_row_map and "current_liabilities" in raw_row_map:
        ca_row = raw_row_map["current_assets"]
        cl_row = raw_row_map["current_liabilities"]
        ws.cell(row=row, column=2, value=f"=B{ca_row}-B{cl_row}")
        ws.cell(row=row, column=3, value=f"=C{ca_row}-C{cl_row}")
        _apply_data_style(ws.cell(row=row, column=2), is_formula=True)
        _apply_data_style(ws.cell(row=row, column=3), is_formula=True)
    else:
        ws.cell(row=row, column=2, value=ratios.working_capital)
        ws.cell(row=row, column=3, value=ratios.working_capital_prior)
    ws.cell(row=row, column=4, value=f"=B{row}-C{row}")
    _apply_data_style(ws.cell(row=row, column=4), is_formula=True)
    ws.cell(row=row, column=5, value="Current Assets - Current Liabilities")
    ws.cell(row=row, column=2).number_format = '#,##0'
    ws.cell(row=row, column=3).number_format = '#,##0'
    ws.cell(row=row, column=4).number_format = '#,##0'
    row += 1

    # Return on Assets
    ws.cell(row=row, column=1, value="Return on Assets")
    if "net_income" in raw_row_map and "total_assets" in raw_row_map:
        ni_row = raw_row_map["net_income"]
        ta_row = raw_row_map["total_assets"]
        ws.cell(row=row, column=2, value=f"=B{ni_row}/B{ta_row}")
        ws.cell(row=row, column=3, value=f"=C{ni_row}/C{ta_row}")
        _apply_data_style(ws.cell(row=row, column=2), is_formula=True)
        _apply_data_style(ws.cell(row=row, column=3), is_formula=True)
    else:
        ws.cell(row=row, column=2, value=ratios.return_on_assets)
        ws.cell(row=row, column=3, value=ratios.return_on_assets_prior)
    ws.cell(row=row, column=4, value=f"=B{row}-C{row}")
    _apply_data_style(ws.cell(row=row, column=4), is_formula=True)
    ws.cell(row=row, column=5, value="Net Income / Total Assets")
    ws.cell(row=row, column=2).number_format = '0.00%'
    ws.cell(row=row, column=3).number_format = '0.00%'
    ws.cell(row=row, column=4).number_format = '+0.00%;-0.00%'
    row += 1

    # Return on Equity
    ws.cell(row=row, column=1, value="Return on Equity")
    if "net_income" in raw_row_map and "stockholders_equity" in raw_row_map:
        ni_row = raw_row_map["net_income"]
        se_row = raw_row_map["stockholders_equity"]
        ws.cell(row=row, column=2, value=f"=B{ni_row}/B{se_row}")
        ws.cell(row=row, column=3, value=f"=C{ni_row}/C{se_row}")
        _apply_data_style(ws.cell(row=row, column=2), is_formula=True)
        _apply_data_style(ws.cell(row=row, column=3), is_formula=True)
    else:
        ws.cell(row=row, column=2, value=ratios.return_on_equity)
        ws.cell(row=row, column=3, value=ratios.return_on_equity_prior)
    ws.cell(row=row, column=4, value=f"=B{row}-C{row}")
    _apply_data_style(ws.cell(row=row, column=4), is_formula=True)
    ws.cell(row=row, column=5, value="Net Income / Stockholders' Equity")
    ws.cell(row=row, column=2).number_format = '0.00%'
    ws.cell(row=row, column=3).number_format = '0.00%'
    ws.cell(row=row, column=4).number_format = '+0.00%;-0.00%'

    # Auto-fit columns
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 40


SECTION_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
SUBTOTAL_FONT = Font(bold=True)
SUBTOTAL_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='medium'),
    bottom=Side(style='thin'),
)


def _create_statement_sheet(
    wb: Workbook,
    session: ExtractionSession,
    sheet_name: str,
    items: list,
):
    """Create a financial statement sheet (Income Statement, Balance Sheet, or Cash Flow)."""
    ws = wb.create_sheet(sheet_name)

    # Headers
    headers = ["Item", "Current Year", "Prior Year", "YoY Delta"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        _apply_header_style(cell)

    row = 2
    last_section = ""
    raw_row_map = {}  # metric_key -> row number for formulas

    for item in items:
        if item.metric_key not in session.raw_values:
            continue

        # Section header
        if item.section and item.section != last_section:
            last_section = item.section
            section_cell = ws.cell(row=row, column=1, value=item.section.upper())
            section_cell.font = Font(bold=True, color="1F4E79", size=10)
            section_cell.fill = SECTION_FILL
            for col in range(2, 5):
                ws.cell(row=row, column=col).fill = SECTION_FILL
            row += 1
        elif not item.section and last_section:
            last_section = ""

        ev = session.raw_values[item.metric_key]
        raw_row_map[item.metric_key] = row

        # Label with indent
        label = ("    " if item.indent_level > 0 else "") + item.display_name
        ws.cell(row=row, column=1, value=label)

        # Values
        ws.cell(row=row, column=2, value=ev.value)
        ws.cell(row=row, column=3, value=ev.value_prior)

        # Delta formula
        ws.cell(row=row, column=4, value=f"=B{row}-C{row}")
        _apply_data_style(ws.cell(row=row, column=4), is_formula=True)

        # Number formats
        ws.cell(row=row, column=2).number_format = '#,##0'
        ws.cell(row=row, column=3).number_format = '#,##0'
        ws.cell(row=row, column=4).number_format = '#,##0'

        # Bold + border for subtotal rows
        if item.is_bold:
            for col in range(1, 5):
                cell = ws.cell(row=row, column=col)
                cell.font = SUBTOTAL_FONT
                if item.is_subtotal:
                    cell.border = SUBTOTAL_BORDER

        # Apply cell borders
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = THIN_BORDER

        row += 1

    # Auto-fit columns
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18


def _create_verification_sheet(wb: Workbook, verification: VerificationResult):
    """Create the Verification sheet showing all checks with pass/fail."""
    ws = wb.create_sheet("Verification")

    # Summary at top
    ws.cell(row=1, column=1, value="Verification Summary")
    ws.cell(row=1, column=1).font = Font(bold=True, size=12, color="1F4E79")

    ws.cell(row=2, column=1, value=f"Passed: {verification.pass_count}")
    ws.cell(row=2, column=1).font = Font(color="006100")
    ws.cell(row=2, column=2, value=f"Failed: {verification.fail_count}")
    ws.cell(row=2, column=2).font = Font(color="9C0006")
    ws.cell(row=2, column=3, value=f"Skipped: {verification.skip_count}")
    ws.cell(row=2, column=3).font = Font(color="808080")

    # Headers
    headers = ["Check", "Year", "Formula", "LHS Value", "RHS Value", "Difference", "Tolerance", "Result", "Severity"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        _apply_header_style(cell)

    # Data rows
    for idx, check in enumerate(verification.checks, 5):
        ws.cell(row=idx, column=1, value=check.description)
        ws.cell(row=idx, column=2, value=check.year)
        ws.cell(row=idx, column=3, value=check.formula)
        ws.cell(row=idx, column=4, value=check.lhs_value)
        ws.cell(row=idx, column=5, value=check.rhs_value)
        ws.cell(row=idx, column=6, value=check.difference)
        ws.cell(row=idx, column=7, value=f"{check.tolerance * 100:.1f}%")

        result_text = "SKIP" if check.skipped else ("PASS" if check.passed else "FAIL")
        result_cell = ws.cell(row=idx, column=8, value=result_text)

        if check.skipped:
            result_cell.font = Font(color="808080")
        elif check.passed:
            result_cell.font = Font(color="006100", bold=True)
            result_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        else:
            result_cell.font = Font(color="9C0006", bold=True)
            result_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        ws.cell(row=idx, column=9, value=check.severity.upper())

        # Number formats
        ws.cell(row=idx, column=4).number_format = '#,##0'
        ws.cell(row=idx, column=5).number_format = '#,##0'
        ws.cell(row=idx, column=6).number_format = '#,##0'

        # Borders
        for col in range(1, 10):
            ws.cell(row=idx, column=col).border = THIN_BORDER

    # Auto-fit columns
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].width = 10


def _create_audit_sheet(wb: Workbook, calculation_steps: list[CalculationStep]):
    """Create the Audit Log sheet with calculation steps."""
    ws = wb.create_sheet("Audit Log")

    # Add note at top
    ws.cell(row=1, column=1, value="Note: This sheet documents the calculation logic. Actual working formulas are in the 'Calculated Metrics' and 'Ratios' sheets.")
    ws.cell(row=1, column=1).font = Font(italic=True, color="666666")
    ws.merge_cells('A1:E1')

    # Headers
    headers = ["Metric", "Formula Description", "Inputs Used", "Result"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        _apply_header_style(cell)

    # Data rows
    for row_idx, step in enumerate(calculation_steps, 4):
        ws.cell(row=row_idx, column=1, value=step.metric)
        ws.cell(row=row_idx, column=2, value=step.formula)

        # Format inputs as readable string
        inputs_str = ", ".join([f"{k}={v:,.0f}" for k, v in step.inputs.items()])
        ws.cell(row=row_idx, column=3, value=inputs_str)

        ws.cell(row=row_idx, column=4, value=step.result)
        ws.cell(row=row_idx, column=4).number_format = '#,##0.00'

        # Apply styles
        for col in range(1, 5):
            ws.cell(row=row_idx, column=col).border = THIN_BORDER

    # Auto-fit columns
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 55
    ws.column_dimensions['C'].width = 70
    ws.column_dimensions['D'].width = 20
